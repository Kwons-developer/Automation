""" from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws. title = "시트 제목"
wb.save("엑셀제목.xlsx")
wb.close() """

""" from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet  # 새로운 sheet 기본 이름으로 생성
ws.title = "시트 제목"
ws.sheet_properties.tabColor = "시트 탭 색"

ws1 = wb.create_sheet("시트 제목")
ws2 = wb.create_sheet("시트 제목")

new_ws = wb["NewSheet"]
print(ws.sheetnames)       # 모든 sheet 이름 출력

# sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "제목"

wb.save("~~") """

""" from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = "~~"

ws["A1"] = 1
ws["B1"] = 2

print(ws["A1"])
print(ws["A1"].value)

print(ws.cell(row=1, column=1).value)

ws.cell(row=1, column=1, value=10)  # ws["A1"] = 10

for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=randint(0, 100))

wb.save("~~") """
""" 
from openpyxl import Workbook
from random import *
import sys  # 한글 보여주는것
import io

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding="utf-8")

wb = Workbook()
ws = wb.active

ws.append(("번호", "수학", "영어"))

for i in range(1, 11):
    ws.append((i, randint(0, 100), randint(0, 100)))

for cell in ws[1]:
print(cell.value)

for cols in ws["B:C"]:
    for cell in cols:
        print(cell.value, end=" ")
    print()


for rows in ws[2:6]:
    for row in rows:
        print(row.value, end=" ")
    print()

for row in ws.iter_rows(max_row=6, max_col=2):
    print(row[0].value, row[1].value)

wb.save("allPractice.xlsx")
 """

""" 
from openpyxl import load_workbook
import sys  # 한글 보여주는것
import io

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding="utf-8")


wb = load_workbook(r"C:\Users\User\Desktop\프로그래밍\셀레니움\모든 연습\allPractice.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    if row[1].value > 60:
        print(row[0].value, "번은 수학 천재")
"""


"""
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart, LineChart

wb = load_workbook("allPractice.xlsx")
ws = wb.active """


""" 
bar_value = Reference(ws, min_row=2, max_row=11, min_col=2, max_col=3)
bar_chart = BarChart()
bar_chart.add_data(bar_value)
ws.add_chart(barchart, "E1")
 """
""" 
line_value = Reference(ws, min_row=2, max_row=11, min_col=2, max_col=3)
line_chart = LineChart()
line_chart.add_data(line_value, titles_froom_data=True)
line_chart.title = "성적표"
line_chart.style = 20
line_chart.y_axis.title = "점수"
line_chart.x_axis.title = "번호"

ws.add_chart(line_chart, "E1")

wb.save("allPractice_chart.xlsx") """

