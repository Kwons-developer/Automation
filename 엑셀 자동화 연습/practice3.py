from openpyxl import Workbook
from random import *
import sys  # 한글 보여주는것
import io

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding="utf-8")

wb = Workbook()
ws = wb.active

ws.append(("학번", "수학", "영어"))

for i in range(1, 11):
    ws.append((i , randint(1, 100), randint(1, 100)))

""" for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row = x, column = y).value, end = " ")
    print() """

""" cols_B = ws["B"]
for cols in cols_B:
    print(cols.value)

col_range = ws["B:C"]
for row in col_range:
    for cols in row:
        print(cols.value, end = " ")
    print() """

""" for rows in ws.iter_rows(min_row = 2):
    if rows[1].value > 70:
        print(rows[0].value, "번 학생은 수학 천재입니다.")

 """

for rows in tuple(ws.rows):
    print(rows[0].value)

wb.save("practice3.xlsx")
    