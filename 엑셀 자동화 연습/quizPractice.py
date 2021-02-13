from openpyxl import Workbook
from openpyxl.styles import Alignment
import sys  # 한글 보여주는것
import io

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding="utf-8")

wb = Workbook()
ws = wb.active

ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))

scores = [
(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20)
]

for score in scores:
    ws.append(score)

for idx, cell in enumerate(ws["D"]):
    if idx == 0:
        continue
    cell.value = 10



ws["H1"] = "총점"
ws["I1"] = "성적"

for idx, score in enumerate(scores, 2):         # 숫자는 idx 값에 들어가는 것으로 행렬 값 주는것임
    sum_val = sum(score[1:]) - score[3] + 10 # 학번 제외한 점수 다 더하고 퀴즈2 값 10점으로 바꼇으니까 현재 퀴즈2 값 빼고 10 더해줌
    ws.cell(row = idx, column = 8).value = "=SUM(B{}:G{})".format(idx,idx)

    grade = None
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"
    
    if score[1] < 5:
        grade = "F"

    ws.cell(row = idx, column = 9).value = grade

wb.save("quizPractice.xlsx")