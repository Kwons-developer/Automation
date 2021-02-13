from openpyxl import load_workbook
from random import *
from openpyxl.chart import BarChart, LineChart, Reference

wb = load_workbook("practive.xlsx")
ws = wb.active
""" 
for row in ws.iter_rows(max_row = 11, min_col = 4):     # iter)rows 로는 값 입력 불가??
    row[2].value = randint(1, 100) """


""" bar_value = Reference(ws, min_row = 2, max_row = 11, min_col = 2, max_col = 3)
bar_chart = BarChart()
bar_chart.add_data(bar_value)
ws.add_chart(bar_chart, "D1") """

line_value = Reference(ws, min_row = 2, max_row = 11, min_col = 2, max_col = 3)  # 참조는 min, max 값 다 초기화 해야함
line_chart = LineChart()
line_chart.add_data(line_value, titles_from_data = True)
line_chart.title = "성적표"
line_chart.style = 20
line_chart.x_axis.title = "번호"
line_chart.y_axis.title = "점수"

ws.add_chart(line_chart, "D1")

wb.save("practice2.xlsx")
