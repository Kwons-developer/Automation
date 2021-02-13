from openpyxl import Workbook
from random import *
import sys  # 한글 보여주는것
import io

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding="utf-8")

wb = Workbook()
ws = wb.active

ws.append(("번호", "수학", "영어"))
for i in range(1, 11):
    ws.append((i, randint(1, 100), randint(1, 100)))

""" for row in tuple(ws.rows):      # 튜플은 배열 인덱스 가짐 ws[1] 같은건 인덱스 아니고 ㄹㅇ 1번째
    print(row[0].value) """

for row in ws.iter_rows(min_row=1, max_row =3, min_col = 1, max_col = 3):
    print(row[0].value, row[1].value, row[2].value)

for row in ws[1]:           # 여기서 1은 인덱스 값 아님
    print(row.value)

#튜플이나 iter_rows, iter_cols 는 모두 인덱스 번호 사용



wb.save("practive.xlsx")