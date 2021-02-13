from openpyxl import Workbook
from random import *

import sys  # 한글 보여주는것
import io

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding="utf-8")


wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11):  # 10개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"]  # 영어 column 정보 만 가져오기
print(col_B)

for cell in col_B:  # 영어 column 값 가져오기
    print(cell.value)


col_range = ws["B:C"]  # 영어, 수학 column 함께 가져오기 , 데이터를 가로로 나열 하려면 end 값이 필요함
for cols in col_range:
    for cell in cols:
        print(cell.value, end=" ")
    print()


row_title = ws[1]  # 1번째 행(row) 만 가져오기
for cell in row_title:
    print(cell.value)


row_range = ws[2:6]  # 2번째 행부터  6번째 행까지 가지고 오기
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()


from openpyxl.utils.cell import coordinate_from_string

row = ws[2 : ws.max_row]  # 2번째 부터 (마지막 행을 모를때 쓰는것임) 마지막 행 까지 다가져옴
for rows in row:
    for cell in rows:
        # print(cell.value, end=" ")
        # print(cell.coordinate, end=" ")       #A2, B2 이런식으로 출력
        xy = coordinate_from_string(cell.coordinate)  # ('A', 2) ('B', 2) 이런식으로 출력
        # print(xy, end=" ")
        print(xy[0], end="")  # xy[0] 은 A,B,C 이고 xy[1] 은 행 번호 이다.
        print(xy[1], end=" ")
    print()


print(tuple(ws.rows))  # 전체 rows 정보 가져오기 한 행씩 가져옴
for row in tuple(ws.rows):  # 많은 행 중에서 0번째 행 값 가져온다
    print(row[0].value)


# 전체 columns 정보 가져오기 한 열씩 가져옴
# print(tuple(ws.columns))
#for col in tuple(ws.columns):  # 많은 열 중에서 0번째 열 값 가져온다.
    #print(col[0].value)

#for row in ws.iter_rows():  # 전체 row
    #print(row)

#for column in ws.iter_cols():  # 전체 col
    #print(column[0].value)

# 2번째 행부터 11번째 행까지, 2번째 열부터 3번째 열까지
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):           #좌우 좌우로 출력
    print(row[0].value, row[1].value)  # 수학, 영어

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):        #상하 상하로 출력
    print(col[0].value, col[1].value)


wb.save("sample.xlsx")