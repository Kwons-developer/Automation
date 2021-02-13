from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()  # 새로운 sheet 기본 이름으로 생성
ws.title = "Mysheet"  # sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff"  # sheet 색 추가

ws1 = wb.create_sheet("YourSheet")  # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet", 2)  # 2번째 index에 sheet 생성

new_ws = wb["NewSheet"]  # Dict 형태로 sheet 에 접근 가능

print(wb.sheetnames)  # 모든 sheet 이름 출력

# sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)  # new_ws sheet 복사해서 target에 저장
target.title = "Copied Sheet"

wb.save("sample.xlsx")