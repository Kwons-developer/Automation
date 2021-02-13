from openpyxl import Workbook
import datetime
wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today()    # 오늘 날짜 정보
ws["A2"] = "=SUM(1, 2, 3)" # 1+2+3 더해진 6값 나옴
ws["A3"] = "=AVERAGE(1, 2, 3)"  # 평균 2나옴
ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"  # 30


ws.column_dimensions["A"].width = 50       # width 는 column 에서만 사용 가능
wb.save("sample_formula.xlsx")