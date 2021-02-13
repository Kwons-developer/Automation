from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# A1 셀에 값 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])  # A1 셀 정보 출력
print(ws["A1"].value)  # A1 셀의 값 출력
print(ws["A10"].value)  # 값이 없을 땐 'NONE' 을 출력

# row = 1, 2, 3 ...
# column = A, B, C .... 각 영어는 숫자로 접근 A 는 1 로 접근
print(ws.cell(row=1, column=1).value)  # ws["A1"].value 값과 똑같음

ws.cell(row=1, column=3, value=10)  # ws["C1"].value =10 과 같음



# 반복문 이용해서 랜덤 숫자 채우기
for x in range(1, 11):  # row에 10개 채움
    for y in range(1, 11):  # 10개 column 채움
        ws.cell(row=x, column=y, value=randint(0, 100))  # 0~100 사이 숫자 값 지정


wb.save("sample.xlsx")
