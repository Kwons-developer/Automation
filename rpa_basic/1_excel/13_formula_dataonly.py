from openpyxl import load_workbook
""" wb = load_workbook("sample_formula.xlsx")
ws = wb.active


# 수식 그대로 가져오고 있음
for row in ws.values:       # values 는 각 데이터 값 가져옴
    for cell in row:
        print(cell.value) """


# 계산된 데이터 가져오고 싶다
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# evaluate 되지 않은 상태의 데이터는 None 라고 뜸
for row in ws.values:
    for cell in row:
        print(cell)

