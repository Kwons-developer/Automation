from openpyxl import Workbook
from openpyxl.styles import Alignment



import sys  # 한글 보여주는것
import io

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding="utf-8")

wb = Workbook()
ws = wb.active


scores = [
    ("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"),
    (1,10,8,5,14,26,12),
    (2,7,3,7,15,24,18),
    (3,9,5,8,8,12,4),
    (4,7,8,7,17,21,18),
    (5,7,8,7,16,25,15),
    (6,3,5,8,8,17,0),
    (7,4,9,10,16,27,18),
    (8,6,6,6,15,19,17),
    (9,10,10,9,19,30,19),
    (10,9,8,8,20,25,20)
]

for score in scores:
    ws.append(score)

for row in ws.iter_rows(min_row=2, max_col=4):
    row[3].value = 10                     # row 에서 열 번호는 0부터

""" # 퀴즈2 점수 10 으로 수정
for idx, cell in enumerate(ws["D"])     # 첫번째 행 문자열은 제외하는것
    if idx == 0:        # 제목인 경우 스킵
        continue
    cell.value = 10 """




ws["H1"] = "총점"
ws["H2"] = "=SUM(B2:G2)"
ws["H3"] = "=SUM(B3:G3)"
ws["H4"] = "=SUM(B4:G4)"
ws["H5"] = "=SUM(B5:G5)"
ws["H6"] = "=SUM(B6:G6)"
ws["H7"] = "=SUM(B7:G7)"
ws["H8"] = "=SUM(B8:G8)"
ws["H9"] = "=SUM(B9:G9)"
ws["H10"] = "=SUM(B10:G10)"
ws["H11"] = "=SUM(B11:G11)"

ws["I1"] = "성적"

ws["H1"] = "총점"
ws["I1"] = "성적"

for idx, score in enumerate(scores, start=2):           # start = 1 하면 idx 값에 1부터 시작하는것임
    sum_val = sum(score[1:]) - score[3] + 10 # 학번 제외한 점수 다 더하고 퀴즈2 값 10점으로 바꼇으니까 현재 퀴즈2 값 빼고 10 더해줌
    ws.cell(row=idx, column = 8).value = "=SUM(B{}:G{})".format(idx, idx)
    
    grade = None # 성적
    if sum_val >= 90:
        grade = "A"
    elif sum_val >= 80:
        grade = "B"
    elif sum_val >= 70:
        grade = "C"
    else:
        grade = "D"

    # 출석 5 미만 학생 F
    if score[1] < 5:
        grade = "F"

    ws.cell(row=idx, column = 9).value = grade  # I열이 성적 정보

    







for row in ws[2: ws.max_row]:
    if int(row[7].value) >= 90:
        row[8].value = "A"
    elif int(row[7].value) >= 80:
        row[8].value = "B"
    elif int(row[7].value) >= 70:
        row[8].value = "C"
    else:
        row[8].value = "D"


for row in ws[2: ws.max_row]:
    if int(row[1].value) < 5:
        row[8].value = "F"



for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal = "center", vertical="center")

wb.save("quiz.xlsx")