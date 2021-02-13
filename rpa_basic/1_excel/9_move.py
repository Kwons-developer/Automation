from openpyxl import load_workbook
from random import *

import sys  # 한글 보여주는것
import io

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding="utf-8")

wb = load_workbook("sample.xlsx")
ws = wb.active

#번호 (국어) 영어 수학 

#ws.move_range("B1:C11", rows=0, cols = 1)         # 국어 자리 놓기위해 B,C열 한칸 이동 rows 는 몇줄 밑으로 내리고 cols 는 좌우 이동

#ws.cell(row = 1, column = 2, value = "국어") 
""" ws["B1"].value = "국어"



for rows in ws.iter_rows(max_row = 11, min_col = 2):
    rows[1].value = randint(0, 100) """

ws.move_range("C1:C11", rows = 0, cols = -1)
    

wb.save("sample_move.xlsx")


